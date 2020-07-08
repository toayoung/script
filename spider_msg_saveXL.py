#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time    : 2020-6-30 下午 3:31
# @Author  : TOA
import jsonpath
import requests
import re
import openpyxl

url = 'https://top.chinaz.com/'

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'}

response = requests.get(url, headers=headers)
response = bytes(response.text, response.encoding).decode('utf-8', 'ignore')
h3_titles = re.findall('<div class="LsHead"><h3 class="h3Tit bg-blue">(.*?)</h3></div>', str(response))
h4_titles = re.findall('<h4 class="LsCeHead"><span class="h4Tit"><i></i>(.*?)</span></h4>', str(response))
msg = re.findall('<li class="LsClist">(.*?)</a></li>', str(response))
titles = h3_titles + h4_titles

# 创建模块命名的sheet页
wb = openpyxl.Workbook(r'C:\Users\Au-forever\Desktop\demo.xlsx')
for name in titles:
    wb.create_sheet(title=name)
wb.save(r'C:\Users\Au-forever\Desktop\demo.xlsx')

# 根据模块添加内容
wb = openpyxl.load_workbook(r'C:\Users\Au-forever\Desktop\demo.xlsx')
for index, sheet in enumerate(wb.sheetnames):

    st = wb[sheet]
    st.cell(row=1, column=1, value='排名')
    st.cell(row=1, column=2, value='网站')
    st.cell(row=1, column=3, value='名称')
    st.cell(row=1, column=4, value='一周变化')
    infos = msg[index]
    info = re.findall('<li class="LsClist">.*?">(.*?)</span>.*?"/(.*?).html">(.*?)</a>.*?col-red03">(.*?)</span>',
                      str(infos))
    for data in info:
        num = int(data[0])
        num += 1
        st.cell(row=num, column=1, value=data[0])
        st.cell(row=num, column=2, value=data[1])
        st.cell(row=num, column=3, value=data[2])
        st.cell(row=num, column=4, value=data[3])

wb.save(r'C:\Users\Au-forever\Desktop\demo.xlsx')
